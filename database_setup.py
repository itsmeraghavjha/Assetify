import os
import re
import openpyxl  # <-- THIS IS THE FIX: Use Excel reader
from app import app, db
from models import User, Distributor, AssetRequest

# --- CONFIGURATION ---
# --- THIS IS THE FIX: Point to your .xlsx file ---
EXCEL_FILE_PATH = 'DB vs EMP Mapping (1).xlsx'
SHEET_NAME = 'DB wise - SE Mapping' # Assuming this is the sheet name
# --- END OF FIX ---

# --- PASSWORD RULES ---
ADMIN_PASSWORD = 'adminpass' # A default password for the main admin account


def load_data_from_excel(file_path, sheet_name):
    """
    Loads data from an Excel .xlsx file and returns a list of dictionaries.
    """
    print(f"Reading data from Excel file: {file_path} (Sheet: {sheet_name})")
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
    except FileNotFoundError:
        print(f"--- ERROR ---")
        print(f"File not found: {file_path}")
        print(f"Please make sure the file '{file_path}' is in the same directory as this script.")
        return None
    except Exception as e:
        print(f"--- ERROR Reading Excel File ---")
        print(f"Could not open file. Is it a valid .xlsx file? Error: {e}")
        return None

    try:
        sheet = workbook[sheet_name]
    except KeyError:
        print(f"--- ERROR ---")
        print(f"A sheet named '{sheet_name}' was not found in the Excel file.")
        print(f"Available sheets: {workbook.sheetnames}")
        return None

    rows = []
    # Read headers from the first row
    headers = [str(cell.value).strip() for cell in sheet[1]]
    
    # Read data from all subsequent rows
    for row in sheet.iter_rows(min_row=2):
        row_data = {}
        for header, cell in zip(headers, row):
            row_data[header] = str(cell.value).strip() if cell.value is not None else ''
        rows.append(row_data)
        
    print(f"Successfully read {len(rows)} data rows from Excel.")
    return rows

def setup_database():
    """
    Clears and seeds the database from the provided Excel file.
    """
    
    # --- THIS IS THE FIX: Call the new Excel function ---
    rows = load_data_from_excel(EXCEL_FILE_PATH, SHEET_NAME)
    # --- END OF FIX ---
    
    if rows is None:
        return 

    with app.app_context():
        # --- STAGE 1: Clear existing data ---
        try:
            print("Clearing existing data...")
            db.session.query(AssetRequest).delete()
            
            for user in User.query.all():
                user.distributor_id = None
            db.session.commit()
            
            for dist in Distributor.query.all():
                dist.se_id = None
                dist.bm_id = None
                dist.rh_id = None
            db.session.commit()
            
            db.session.query(User).delete()
            db.session.query(Distributor).delete()
            db.session.commit()
            print("Existing data cleared.")
        except Exception as e:
            db.session.rollback()
            print(f"ERROR clearing data: {e}")
            return

        # --- STAGE 2: Create all User accounts (SE, BM, RH, Admin) ---
        print("Finding and creating all unique User accounts...")
        try:
            unique_ses = {}
            unique_bms = {}
            unique_rhs = {}
            
            for row in rows:
                se_code = row.get('SE Emp Code', '').strip()
                se_name = row.get('SE Name', '').strip()
                
                if (se_code == '0' or not se_code or se_code == 'None') and se_name:
                    se_code = "vacant_" + re.sub(r'[^a-zA-Z0-9]', '_', se_name).lower()

                if se_code and se_code not in unique_ses:
                    unique_ses[se_code] = {
                        'name': se_name,
                        'email': None, 
                        'role': 'SE',
                        'so': row.get('SO', '').strip(),
                        'password': se_code # Rule: SE Pass = SE Emp Code
                    }
                
                bm_code = row.get('BM Emp Code', '').strip()
                if bm_code and bm_code not in unique_bms:
                    unique_bms[bm_code] = {
                        'code': bm_code, 
                        'password': bm_code, # Rule: BM Pass = BM Emp Code
                        'name': row.get('BM', '').strip(),
                        'email': row.get('BM Mail ID', '').strip().lower(),
                        'role': 'BM'
                    }
                    
                rh_code = row.get('RH Emp Code', '').strip()
                if rh_code and rh_code not in unique_rhs:
                    unique_rhs[rh_code] = {
                        'code': rh_code, 
                        'password': rh_code, # Rule: RH Pass = RH Emp Code
                        'name': row.get('RH', '').strip(),
                        'email': row.get('RH Mail ID', '').strip().lower(),
                        'role': 'RH'
                    }
            
            user_map_by_emp_code = {} 
            all_users_to_create = []

            admin_user = User(
                employee_code='admin', 
                name='Admin User',
                email='admin@example.com',
                role='Admin'
            )
            admin_user.set_password(ADMIN_PASSWORD)
            all_users_to_create.append(admin_user)
            user_map_by_emp_code[admin_user.employee_code] = admin_user
            
            # Process SEs
            for code, data in unique_ses.items():
                if code in user_map_by_emp_code: continue
                user = User(
                    employee_code=code,
                    name=data['name'],
                    email=data['email'],
                    role=data['role'],
                    so=data['so']
                )
                user.set_password(data['password'])
                all_users_to_create.append(user)
                user_map_by_emp_code[code] = user

            # Process BMs
            for code, data in unique_bms.items():
                if code in user_map_by_emp_code:
                    print(f"  [WARN] User code {code} (BM) already exists. Skipping.")
                    continue
                user = User(
                    employee_code=code,
                    name=data['name'],
                    email=data['email'],
                    role=data['role']
                )
                user.set_password(data['password'])
                all_users_to_create.append(user)
                user_map_by_emp_code[code] = user
                
            # Process RHs
            for code, data in unique_rhs.items():
                if code in user_map_by_emp_code:
                    print(f"  [WARN] User code {code} (RH) already exists. Skipping.")
                    continue
                user = User(
                    employee_code=code,
                    name=data['name'],
                    email=data['email'],
                    role=data['role']
                )
                user.set_password(data['password'])
                all_users_to_create.append(user)
                user_map_by_emp_code[code] = user

            db.session.add_all(all_users_to_create)
            db.session.commit()
            print(f"Successfully created {len(all_users_to_create)} unique users (Admin, SE, BM, RH).")

        except Exception as e:
            db.session.rollback()
            print(f"--- ERROR Creating Users ---")
            print(f"An error occurred: {e}")
            import traceback
            traceback.print_exc()
            return
            
        # --- STAGE 3: Create Distributors and DB Users ---
        print("Creating and linking distributors...")
        try:
            distributors_to_create = {}
            db_users_to_create = []
            unique_dist_names = set()

            for row in rows:
                dist_code = row.get('Distributor Code', '').strip()
                dist_name = row.get('Distributor Name', 'N/A').strip()
                
                if not dist_code or dist_code == 'None' or dist_code in distributors_to_create:
                    continue 
                
                if dist_name in unique_dist_names:
                    print(f"  [WARN] Distributor name '{dist_name}' (Code: {dist_code}) already exists. Skipping duplicate name.")
                    continue 
                
                se_code = row.get('SE Emp Code', '').strip()
                if (se_code == '0' or not se_code or se_code == 'None'):
                    se_code = "vacant_" + re.sub(r'[^a-zA-Z0-9]', '_', row.get('SE Name', '')).lower()
                se_user = user_map_by_emp_code.get(se_code)
                
                bm_user = user_map_by_emp_code.get(row.get('BM Emp Code', '').strip())
                rh_user = user_map_by_emp_code.get(row.get('RH Emp Code', '').strip())

                dist = Distributor(
                    code=dist_code,
                    name=dist_name,
                    city=row.get('Distributor Town', '').strip(),
                    state=None,
                    se_id=se_user.id if se_user else None,
                    bm_id=bm_user.id if bm_user else None,
                    rh_id=rh_user.id if rh_user else None
                )
                distributors_to_create[dist_code] = dist
                unique_dist_names.add(dist_name) 
                db.session.add(dist)

            db.session.commit()
            print(f"Successfully created {len(distributors_to_create)} distributors.")
            
            print("Creating Distributor (DB) user accounts...")
            for dist_code, dist_obj in distributors_to_create.items():
                if dist_code in user_map_by_emp_code:
                    print(f"  [WARN] User with code '{dist_code}' already exists (likely an SE). Skipping DB user creation.")
                    continue

                db_user = User(
                    employee_code=dist_code, # Rule: DB Login = Dist Code
                    name=f"{dist_obj.name} (DB)",
                    role='DB',
                    email=None,
                    distributor_id=dist_obj.id
                )
                db_user.set_password(dist_code) # Rule: DB Pass = Dist Code
                db_users_to_create.append(db_user)
                user_map_by_emp_code[dist_code] = db_user
            
            db.session.add_all(db_users_to_create)
            db.session.commit()
            print(f"Successfully created {len(db_users_to_create)} Distributor (DB) user accounts.")
            
        except Exception as e:
            db.session.rollback()
            print(f"--- ERROR Creating Distributors or DB Users ---")
            print(f"An error occurred: {e}")
            import traceback
            traceback.print_exc()
            return

        # --- STAGE 4: Create Test Users ---
        print("Creating test users...")
        try:
            test_users = []
            test_distributors = []

            # Test Admins
            test_admin_1 = User(employee_code='testadmin1', name='Test Admin 1', role='Admin', email='admin1@test.com')
            test_admin_1.set_password('testadmin1')
            test_admin_2 = User(employee_code='testadmin2', name='Test Admin 2', role='Admin', email='admin2@test.com')
            test_admin_2.set_password('testadmin2')
            test_users.extend([test_admin_1, test_admin_2])

            # Test SEs
            test_se_1 = User(employee_code='testse1', name='Test SE 1', role='SE', so='TEST-SO')
            test_se_1.set_password('testse1')
            test_se_2 = User(employee_code='testse2', name='Test SE 2', role='SE', so='TEST-SO')
            test_se_2.set_password('testse2')
            test_users.extend([test_se_1, test_se_2])

            # Test BMs
            test_bm_1 = User(employee_code='testbm1', name='Test BM 1', role='BM', email='bm1@test.com')
            test_bm_1.set_password('testbm1')
            test_bm_2 = User(employee_code='testbm2', name='Test BM 2', role='BM', email='bm2@test.com')
            test_bm_2.set_password('testbm2')
            test_users.extend([test_bm_1, test_bm_2])

            # Test RHs
            test_rh_1 = User(employee_code='testrh1', name='Test RH 1', role='RH', email='rh1@test.com')
            test_rh_1.set_password('testrh1')
            test_rh_2 = User(employee_code='testrh2', name='Test RH 2', role='RH', email='rh2@test.com')
            test_rh_2.set_password('testrh2')
            test_users.extend([test_rh_1, test_rh_2])
            
            db.session.add_all(test_users)
            db.session.commit() # Commit users first to get IDs

            # Test Distributors (linked to test managers)
            test_dist_1 = Distributor(code='TESTDIST1', name='Test Distributor 1', city='Test City', se_id=test_se_1.id, bm_id=test_bm_1.id, rh_id=test_rh_1.id)
            test_dist_2 = Distributor(code='TESTDIST2', name='Test Distributor 2', city='Test City', se_id=test_se_2.id, bm_id=test_bm_2.id, rh_id=test_rh_2.id)
            test_distributors.extend([test_dist_1, test_dist_2])
            
            db.session.add_all(test_distributors)
            db.session.commit() # Commit distributors to get IDs

            # Test DB Users (linked to test distributors)
            test_db_1 = User(employee_code='testdb1', name='Test DB User 1', role='DB', distributor_id=test_dist_1.id)
            test_db_1.set_password('testdb1')
            test_db_2 = User(employee_code='testdb2', name='Test DB User 2', role='DB', distributor_id=test_dist_2.id)
            test_db_2.set_password('testdb2')

            db.session.add_all([test_db_1, test_db_2])
            db.session.commit()
            
            print(f"Successfully created {len(test_users) + 2} test users and {len(test_distributors)} test distributors.")

        except Exception as e:
            db.session.rollback()
            print(f"--- ERROR Creating Test Users ---")
            print(f"An error occurred: {e}")
            import traceback
            traceback.print_exc()

        # --- STAGE 5: Final Summary ---
        print("\n" + "="*60)
        print("DATABASE SEEDED SUCCESSFULLY WITH REAL & TEST DATA!")
        print("="*60)
        print("\n--- REAL DATA LOGIN CREDENTIALS ---")
        print("\nADMIN:")
        print(f"  User: admin / Pass: {ADMIN_PASSWORD}")
        
        print("\nBRANCH MANAGERS (BMs):")
        print("  (Employee Code is Password)")
        for code, data in unique_bms.items(): 
            print(f"  - {code} ({data['name']})")

        print("\nREGIONAL HEADS (RHs):")
        print("  (Employee Code is Password)")
        for code, data in unique_rhs.items(): 
            print(f"  - {code} ({data['name']})")

        print("\nSALES EXECUTIVES (SEs):")
        print("  (Employee Code is Password)")
        for code, data in unique_ses.items(): 
            print(f"  - {code} ({data['name']})")

        print("\nDISTRIBUTORS (DBs):")
        print("  (Distributor Code is Employee Code AND Password)")
        for code, data in distributors_to_create.items(): 
            if code not in unique_ses:
                print(f"  - {code} ({data.name})")
        
        print("\n" + "="*60)
        print("\n--- TEST USER LOGIN CREDENTIALS ---")
        print("  (Username is Password for all test accounts)")
        print("\n  TEST ADMINS:")
        print("  - testadmin1 / testadmin1")
        print("  - testadmin2 / testadmin2")
        print("\n  TEST SEs:")
        print("  - testse1 / testse1")
        print("  - testse2 / testse2")
        print("\n  TEST BMs:")
        print("  - testbm1 / testbm1")
        print("  - testbm2 / testbm2")
        print("\n  TEST RHs:")
        print("  - testrh1 / testrh1")
        print("  - testrh2 / testrh2")
        print("\n  TEST DBs (Linked to Test Distributors):")
        print("  - testdb1 / testdb1")
        print("  - testdb2 / testdb2")
        
        print("="*60)

if __name__ == '__main__':
    setup_database()